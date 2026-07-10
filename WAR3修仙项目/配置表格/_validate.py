import openpyxl, os

base = r'E:\thoth\横向\WAR3修仙项目\配置表格'

# Load enums - full_name in column B (2)
wb = openpyxl.load_workbook(os.path.join(base, '__enums__.xlsx'), data_only=True)
ws = wb['Sheet1']
enums = {}
current_enum = None
for rn in range(4, ws.max_row + 1):
    ename = ws.cell(row=rn, column=2).value  # B = full_name
    if ename is not None:
        current_enum = str(ename)
        enums[current_enum] = []
    name_val = ws.cell(row=rn, column=8).value  # H = name
    alias_val = ws.cell(row=rn, column=9).value  # I = alias
    value_val = ws.cell(row=rn, column=10).value  # J = value
    if name_val is not None and current_enum:
        enums[current_enum].append((str(name_val), str(alias_val) if alias_val else '', value_val))

print(f"Enums: {list(enums.keys())}")
for k, v in enums.items():
    print(f"  {k}: {len(v)} items")
    for item in v[:3]:
        print(f"    {item}")
    if len(v) > 3:
        print(f"    ... +{len(v)-3} more")

# Now validate quality
iq = enums.get('ItemQuality', [])
iq_names = [e[0] for e in iq]  # White, Green...
iq_aliases = [e[1] for e in iq]  # 白色, 绿色...
print(f"\nItemQuality names={iq_names}, aliases={iq_aliases}")

fp = os.path.join(base, 'W物品表.xlsx')
wb2 = openpyxl.load_workbook(fp, data_only=True)
ws2 = wb2['item_config']
invalid_color = []
for rn in range(5, ws2.max_row + 1):
    a1 = ws2.cell(row=rn, column=1).value
    if a1 and str(a1).startswith('##'):
        continue
    q = ws2.cell(row=rn, column=3).value
    if q:
        qs = str(q)
        if qs not in iq_names and qs not in iq_aliases:
            invalid_color.append((rn, qs))

if invalid_color:
    print(f"\n❌ Invalid quality values in item_config: {invalid_color[:10]}")
else:
    print(f"\n✅ All quality values in item_config valid")

# Monster dignity
md = enums.get('MonsterDignity', [])
md_names = [e[0] for e in md]
fp = os.path.join(base, 'G怪物表.xlsx')
wb3 = openpyxl.load_workbook(fp, data_only=True)
ws3 = wb3['monster_config']
invalid_dignity = []
for rn in range(5, ws3.max_row + 1):
    a1 = ws3.cell(row=rn, column=1).value
    if a1 and str(a1).startswith('##'):
        continue
    d = ws3.cell(row=rn, column=10).value
    if d and str(d) not in md_names:
        invalid_dignity.append((rn, d))
if invalid_dignity:
    print(f"❌ Invalid dignity in monster_config: {invalid_dignity}")
else:
    print("✅ All dignity values valid")

# EquipPart
ep = enums.get('EquipPart', [])
ep_names = [e[0] for e in ep]
fp = os.path.join(base, 'Z装备.xlsx')
wb4 = openpyxl.load_workbook(fp, data_only=True)
ws4 = wb4['equip_config']
invalid_ep = []
for rn in range(5, ws4.max_row + 1):
    a1 = ws4.cell(row=rn, column=1).value
    if a1 and str(a1).startswith('##'):
        continue
    e = ws4.cell(row=rn, column=7).value
    if e and str(e) not in ep_names:
        invalid_ep.append((rn, e))
if invalid_ep:
    print(f"❌ Invalid EquipPart: {invalid_ep}")
else:
    print("✅ All EquipPart values valid")

# EquipRank
er = enums.get('EquipRank', [])
er_names = [e[0] for e in er]
invalid_er = []
for rn in range(5, ws4.max_row + 1):
    a1 = ws4.cell(row=rn, column=1).value
    if a1 and str(a1).startswith('##'):
        continue
    e = ws4.cell(row=rn, column=6).value
    if e and str(e) not in er_names:
        invalid_er.append((rn, e))
if invalid_er:
    print(f"❌ Invalid EquipRank: {invalid_er}")
else:
    print("✅ All EquipRank values valid")

# DropType
dt = enums.get('DropType', [])
dt_names = [e[0] for e in dt]
fp = os.path.join(base, 'D掉落方案.xlsx')
wb5 = openpyxl.load_workbook(fp, data_only=True)
ws5 = wb5['drop_config']
invalid_dt = []
for rn in range(5, ws5.max_row + 1):
    a1 = ws5.cell(row=rn, column=1).value
    if a1 and str(a1).startswith('##'):
        continue
    d = ws5.cell(row=rn, column=5).value
    if d and str(d) not in dt_names:
        invalid_dt.append((rn, d))
if invalid_dt:
    print(f"❌ Invalid DropType: {invalid_dt[:10]}")
else:
    print("✅ All DropType values valid")

# ItemDisplayType
idt = enums.get('ItemDisplayType', [])
idt_names = [e[0] for e in idt]
invalid_idt = []
for rn in range(5, ws2.max_row + 1):
    a1 = ws2.cell(row=rn, column=1).value
    if a1 and str(a1).startswith('##'):
        continue
    d = ws2.cell(row=rn, column=8).value
    if d and str(d) not in idt_names:
        invalid_idt.append((rn, d))
if invalid_idt:
    print(f"❌ Invalid DisplayType: {invalid_idt[:10]}")
else:
    print("✅ All DisplayType values valid")

# AttrKey in monster_config attrs
ak = enums.get('AttrKey', [])
ak_names = [e[0] for e in ak]
fp = os.path.join(base, 'G怪物表.xlsx')
ws6 = wb3['monster_config']
invalid_ak = []
for rn in range(5, ws6.max_row + 1):
    a1 = ws6.cell(row=rn, column=1).value
    if a1 and str(a1).startswith('##'):
        continue
    for pair in range(14):
        key_col = 35 + pair * 2
        key = ws6.cell(row=rn, column=key_col).value
        if key and str(key) not in ak_names:
            invalid_ak.append((rn, key))
if invalid_ak:
    print(f"❌ Invalid AttrKey in monster_config: {invalid_ak[:10]}")
else:
    print("✅ All monster_config AttrKey values valid")

# AttrKey in equip_config attrs
invalid_ak2 = []
for rn in range(5, ws4.max_row + 1):
    a1 = ws4.cell(row=rn, column=1).value
    if a1 and str(a1).startswith('##'):
        continue
    for pair in range(15):
        key_col = 38 + pair * 2
        key = ws4.cell(row=rn, column=key_col).value
        if key and str(key) not in ak_names:
            invalid_ak2.append((rn, key))
if invalid_ak2:
    print(f"❌ Invalid AttrKey in equip_config: {invalid_ak2[:10]}")
else:
    print("✅ All equip_config AttrKey values valid")

# ConfigScene commented rows
fp = os.path.join(base, 'C场景区域配置.xlsx')
wb7 = openpyxl.load_workbook(fp, data_only=True)
ws7 = wb7['ConfigScene']
commented = []
active = []
for rn in range(5, ws7.max_row + 1):
    a1 = ws7.cell(row=rn, column=1).value
    vals = [ws7.cell(row=rn, column=c).value for c in range(1, 5)]
    if a1 and str(a1).startswith('##'):
        commented.append(rn)
    elif any(v is not None for v in vals):
        active.append(rn)
print(f"\n⚠️ ConfigScene: {len(active)} active, {len(commented)} commented rows")
if commented and active == 1:
    print(f"   Only row 5 (scene_id=1) is active. All others are ##-commented.")

# region_config commented rows
ws8 = wb7['region_config']
cregions = []
aregions = []
for rn in range(5, ws8.max_row + 1):
    a1 = ws8.cell(row=rn, column=1).value
    c1 = ws8.cell(row=rn, column=1).value
    c2 = ws8.cell(row=rn, column=2).value
    if a1 and str(a1).startswith('##'):
        cregions.append(rn)
    elif c2:
        aregions.append(rn)
print(f"⚠️ region_config: {len(aregions)} active, {len(cregions)} commented rows")

# Monster_config: check for #N/A
ws9 = wb3['monster_config']
na_issues = []
for rn in range(5, ws9.max_row + 1):
    a1 = ws9.cell(row=rn, column=1).value
    if a1 and str(a1).startswith('##'):
        continue
    for c in [1, 2, 5, 6, 10]:
        v = ws9.cell(row=rn, column=c).value
        if v and str(v) == '#N/A':
            na_issues.append((rn, c))
if na_issues:
    print(f"\n❌ monster_config has #N/A values: {na_issues}")
else:
    print("\n✅ No #N/A values in monster_config")

# Spawn_config: check monster_id references
fp = os.path.join(base, 'G怪物刷新配置.xlsx')
wb10 = openpyxl.load_workbook(fp, data_only=True)
ws10 = wb10['spawn_config']
# Load all monster config IDs
monster_ids = set()
for rn in range(5, ws9.max_row + 1):
    a1 = ws9.cell(row=rn, column=1).value
    if a1 and str(a1).startswith('##'):
        continue
    cid = ws9.cell(row=rn, column=1).value
    if cid:
        monster_ids.add(int(cid))

missing_spawn = []
for rn in range(5, ws10.max_row + 1):
    a1 = ws10.cell(row=rn, column=1).value
    if a1 and str(a1).startswith('##'):
        continue
    for pair in range(10):
        id_col = 8 + pair * 2  # H = Monster1ID
        mid = ws10.cell(row=rn, column=id_col).value
        if mid is not None and str(mid).strip().isdigit():
            mid_int = int(mid)
            if mid_int not in monster_ids and mid_int > 0:
                missing_spawn.append((rn, mid_int))
if missing_spawn:
    unique_missing = list(set(m[1] for m in missing_spawn))
    print(f"❌ spawn_config references {len(missing_spawn)} unknown monster IDs: {unique_missing[:20]}")
else:
    print(f"✅ All spawn_config monster refs valid ({len(monster_ids)} known monsters)")

print("\n=== DONE ===")
